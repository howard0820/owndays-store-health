"""
analyze_decline.py — Deep-dive analysis to find the clearest angles showing
inventory health deterioration from Feb → Apr 2026.

Reads the 5 downloaded StoreHealth files, computes multiple metrics,
and outputs a comprehensive Excel + console summary.

Usage:
  python analyze_decline.py
  python analyze_decline.py --skip 2026-01-01   (skip a date if file missing)
"""

import os, sys, argparse
import pandas as pd
import numpy as np
from datetime import datetime, date, timedelta

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE_DIR)

from store_health_core import (
    parse_inventory_file, classify_category, compute_metrics,
    compute_rankings, is_retail_store, is_warehouse,
    fetch_live_store_list, filter_by_live_list,
    OOS_THRESHOLD, LOW_STOCK_DOH, BRAND_DISPLAYABLE_MIN
)

TARGET_DATES = [
    date(2026, 1, 1),
    date(2026, 2, 1),
    date(2026, 3, 1),
    date(2026, 4, 1),
    date(2026, 4, 20),
]
SALES_DAYS = 30
TOP_N_TIERS = [50, 100, 200, 300]


def load_and_analyze(bd, filepath):
    """Full analysis pipeline for one date. Returns dict of metrics + df."""
    print(f"\n{'─'*50}")
    print(f"  Loading {bd}...")

    df = parse_inventory_file(filepath)
    df = classify_category(df)
    df_main = df[df['category'].isin(['光學眼鏡', '太陽眼鏡'])].copy()

    retail = [s for s in df_main['store_name'].unique() if is_retail_store(s)]
    wh = [s for s in df_main['store_name'].unique() if is_warehouse(s)]

    live, not_live = fetch_live_store_list()
    if live:
        retail, _ = filter_by_live_list(retail, live, not_live_names=not_live)

    df_r = df_main[df_main['store_name'].isin(retail + wh)].copy()
    df_r = compute_metrics(df_r, SALES_DAYS)
    df_r = compute_rankings(df_r, SALES_DAYS, category_groups={'frames': ['光學眼鏡', '太陽眼鏡']})

    df_frames = df_r[
        df_r['category'].isin(['光學眼鏡', '太陽眼鏡']) &
        df_r['store_name'].isin(retail)
    ].copy()

    n_stores = len(retail)
    total = len(df_frames)

    # ── Angle 1: Overall rates ──
    oos = df_frames['is_stockout'].sum()
    low = df_frames['is_low_stock'].sum()

    # ── Angle 2: Top-N rates + displayable store count ──
    tier_data = {}
    for n in TOP_N_TIERS:
        t = df_frames[df_frames['national_rank'] <= n]
        ct = len(t)
        t_oos = t['is_stockout'].sum()
        t_low = t['is_low_stock'].sum()
        valid_doh = t[t['doh_days'] < 9999]['doh_months']

        # Displayable stores: per SKU, count stores with qty >= 2
        disp = t[t['inventory'] >= BRAND_DISPLAYABLE_MIN].groupby('PLU')['store_name'].nunique()
        avg_disp_stores = disp.mean() if len(disp) else 0
        # Also compute: avg displayable store ratio (out of total stores)
        avg_disp_ratio = avg_disp_stores / n_stores * 100 if n_stores else 0

        tier_data[n] = {
            'total': ct,
            'stockout_pct': t_oos / ct * 100 if ct else 0,
            'low_pct': t_low / ct * 100 if ct else 0,
            'shortage_pct': (t_oos + t_low) / ct * 100 if ct else 0,
            'avg_doh': valid_doh.mean() if len(valid_doh) else 0,
            'median_doh': valid_doh.median() if len(valid_doh) else 0,
            'avg_disp_stores': round(avg_disp_stores, 1),
            'avg_disp_ratio': round(avg_disp_ratio, 1),
        }

    # ── Angle 3: Store-level crisis count ──
    store_crisis = {'50pct': 0, '60pct': 0, '70pct': 0, '80pct': 0}
    store_pcts = []
    for store in retail:
        sf = df_frames[df_frames['store_name'] == store].sort_values('national_rank').head(50)
        if len(sf) == 0:
            continue
        pct = (sf['is_stockout'].sum() + sf['is_low_stock'].sum()) / len(sf) * 100
        store_pcts.append(pct)
        if pct > 50: store_crisis['50pct'] += 1
        if pct > 60: store_crisis['60pct'] += 1
        if pct > 70: store_crisis['70pct'] += 1
        if pct > 80: store_crisis['80pct'] += 1

    # ── Angle 4: Brand displayability ──
    brand_data = {}
    for store in retail:
        sf = df_frames[(df_frames['store_name'] == store) & (df_frames['inventory'] >= BRAND_DISPLAYABLE_MIN)]
        if 'ブランド' in sf.columns:
            n_brands = sf['ブランド'].nunique()
            brand_data[store] = n_brands
    avg_brands = np.mean(list(brand_data.values())) if brand_data else 0

    # ── Angle 5: Avg DOH distribution ──
    valid_doh_all = df_frames[df_frames['doh_days'] < 9999]['doh_months']
    doh_under_1m = (valid_doh_all < 1).sum() / len(valid_doh_all) * 100 if len(valid_doh_all) else 0
    doh_under_2m = (valid_doh_all < 2).sum() / len(valid_doh_all) * 100 if len(valid_doh_all) else 0

    # ── Angle 6: Zero-stock SKU count (inventory=0, not even display piece) ──
    zero_stock = (df_frames['inventory'] == 0).sum()
    zero_stock_pct = zero_stock / total * 100 if total else 0

    # ── Angle 7: Top 50 median WOC (weeks of cover) ──
    top50 = df_frames[df_frames['national_rank'] <= 50]
    top50_valid = top50[top50['doh_days'] < 9999]
    top50_median_woc = top50_valid['doh_weeks'].median() if len(top50_valid) else 0

    # ── Angle 8: Per-store avg top50 shortage ──
    avg_store_top50_shortage = np.mean(store_pcts) if store_pcts else 0
    worst_store_pct = max(store_pcts) if store_pcts else 0
    best_store_pct = min(store_pcts) if store_pcts else 0

    result = {
        'date': bd,
        'n_stores': n_stores,
        'total_rows': total,
        'unique_skus': df_frames['PLU'].nunique(),
        # Overall
        'overall_stockout_pct': oos / total * 100 if total else 0,
        'overall_low_pct': low / total * 100 if total else 0,
        'overall_shortage_pct': (oos + low) / total * 100 if total else 0,
        # Top-N
        'tier_data': tier_data,
        # Store crisis
        'stores_over_50pct': store_crisis['50pct'],
        'stores_over_60pct': store_crisis['60pct'],
        'stores_over_70pct': store_crisis['70pct'],
        'stores_over_80pct': store_crisis['80pct'],
        # Brand
        'avg_displayable_brands': round(avg_brands, 1),
        # DOH distribution
        'doh_under_1m_pct': doh_under_1m,
        'doh_under_2m_pct': doh_under_2m,
        # Zero stock
        'zero_stock_pct': zero_stock_pct,
        # Top50 WOC
        'top50_median_woc_weeks': top50_median_woc,
        # Store-level
        'avg_store_top50_shortage': avg_store_top50_shortage,
        'worst_store_pct': worst_store_pct,
        'best_store_pct': best_store_pct,
    }
    return result, df_frames


def print_summary(all_results):
    """Print console summary highlighting the strongest decline signals."""
    dates = [r['date'] for r in all_results]

    print(f"\n{'='*70}")
    print(f"  INVENTORY HEALTH DECLINE ANALYSIS — Feb → Apr 2026")
    print(f"{'='*70}")

    # Find Feb and latest
    feb = next((r for r in all_results if r['date'].month == 2), None)
    latest = all_results[-1]

    if not feb:
        print("  WARNING: No Feb data found")
        feb = all_results[0]

    # Rank angles by magnitude of change
    angles = []

    # 1. Top 50 shortage
    feb_t50 = feb['tier_data'][50]['shortage_pct']
    lat_t50 = latest['tier_data'][50]['shortage_pct']
    delta = lat_t50 - feb_t50
    angles.append(('Top 50 庫存不足率', feb_t50, lat_t50, delta, 'pct'))

    # 2. Top 100 shortage
    feb_t100 = feb['tier_data'][100]['shortage_pct']
    lat_t100 = latest['tier_data'][100]['shortage_pct']
    delta100 = lat_t100 - feb_t100
    angles.append(('Top 100 庫存不足率', feb_t100, lat_t100, delta100, 'pct'))

    # 3. Top 50 stockout only
    feb_oos = feb['tier_data'][50]['stockout_pct']
    lat_oos = latest['tier_data'][50]['stockout_pct']
    angles.append(('Top 50 缺貨率 (qty≤1)', feb_oos, lat_oos, lat_oos - feb_oos, 'pct'))

    # 4. Stores with >50% shortage
    angles.append(('門市 Top50不足率 >50%', feb['stores_over_50pct'], latest['stores_over_50pct'],
                    latest['stores_over_50pct'] - feb['stores_over_50pct'], 'count'))

    # 5. Stores with >70% shortage
    angles.append(('門市 Top50不足率 >70%', feb['stores_over_70pct'], latest['stores_over_70pct'],
                    latest['stores_over_70pct'] - feb['stores_over_70pct'], 'count'))

    # 6. Avg DOH top 50
    feb_doh = feb['tier_data'][50]['avg_doh']
    lat_doh = latest['tier_data'][50]['avg_doh']
    angles.append(('Top 50 Avg DOH (月)', feb_doh, lat_doh, lat_doh - feb_doh, 'num_neg'))

    # 7. Median WOC top 50
    angles.append(('Top 50 Median WOC (週)', feb['top50_median_woc_weeks'],
                    latest['top50_median_woc_weeks'],
                    latest['top50_median_woc_weeks'] - feb['top50_median_woc_weeks'], 'num_neg'))

    # 8. Avg store top50 shortage
    angles.append(('門市平均 Top50 不足率', feb['avg_store_top50_shortage'],
                    latest['avg_store_top50_shortage'],
                    latest['avg_store_top50_shortage'] - feb['avg_store_top50_shortage'], 'pct'))

    # 9. DOH <1m %
    angles.append(('DOH<1月 的 SKU 占比', feb['doh_under_1m_pct'], latest['doh_under_1m_pct'],
                    latest['doh_under_1m_pct'] - feb['doh_under_1m_pct'], 'pct'))

    # 10. Brand displayability
    angles.append(('平均可展示品牌數', feb['avg_displayable_brands'], latest['avg_displayable_brands'],
                    latest['avg_displayable_brands'] - feb['avg_displayable_brands'], 'num_neg'))

    # 11. Worst store
    angles.append(('最差門市 Top50 不足率', feb['worst_store_pct'], latest['worst_store_pct'],
                    latest['worst_store_pct'] - feb['worst_store_pct'], 'pct'))

    # 12. Top SKU displayable store count
    for n in [50, 100]:
        feb_ds = feb['tier_data'][n]['avg_disp_stores']
        lat_ds = latest['tier_data'][n]['avg_disp_stores']
        angles.append((f'Top {n} 平均可展示間數', feb_ds, lat_ds, lat_ds - feb_ds, 'num_neg'))
        feb_dr = feb['tier_data'][n]['avg_disp_ratio']
        lat_dr = latest['tier_data'][n]['avg_disp_ratio']
        angles.append((f'Top {n} 可展示率(間/總店)', feb_dr, lat_dr, lat_dr - feb_dr, 'pct_neg'))

    # 16. Zero stock
    angles.append(('完全零庫存 占比', feb['zero_stock_pct'], latest['zero_stock_pct'],
                    latest['zero_stock_pct'] - feb['zero_stock_pct'], 'pct'))

    # Sort by absolute delta (strongest signal first)
    angles.sort(key=lambda x: abs(x[3]), reverse=True)

    print(f"\n  {'指標':<28} {'Feb':>10} {'Latest':>10} {'Δ':>10}  方向")
    print(f"  {'─'*28} {'─'*10} {'─'*10} {'─'*10}  ────")
    for name, feb_v, lat_v, delta, fmt in angles:
        if fmt == 'pct':
            arrow = '↑惡化' if delta > 0 else '↓改善'
            print(f"  {name:<28} {feb_v:>9.1f}% {lat_v:>9.1f}% {delta:>+9.1f}pp {arrow}")
        elif fmt == 'pct_neg':
            arrow = '↓惡化' if delta < 0 else '↑改善'
            print(f"  {name:<28} {feb_v:>9.1f}% {lat_v:>9.1f}% {delta:>+9.1f}pp {arrow}")
        elif fmt == 'count':
            arrow = '↑惡化' if delta > 0 else '↓改善'
            print(f"  {name:<28} {int(feb_v):>9}家 {int(lat_v):>9}家 {delta:>+9.0f}家  {arrow}")
        elif fmt == 'num_neg':
            arrow = '↓惡化' if delta < 0 else '↑改善'
            print(f"  {name:<28} {feb_v:>9.2f}  {lat_v:>9.2f}  {delta:>+9.2f}   {arrow}")

    # Full timeline for top angles
    print(f"\n{'─'*70}")
    print(f"  TOP 3 最明顯惡化指標 — 完整時間線")
    print(f"{'─'*70}")

    for name, _, _, _, fmt in angles[:3]:
        print(f"\n  {name}:")
        for r in all_results:
            d = r['date']
            if name.startswith('Top 50 庫存不足率'):
                v = r['tier_data'][50]['shortage_pct']
                print(f"    {d}:  {v:.1f}%")
            elif name.startswith('Top 100 庫存不足率'):
                v = r['tier_data'][100]['shortage_pct']
                print(f"    {d}:  {v:.1f}%")
            elif name.startswith('Top 50 缺貨率'):
                v = r['tier_data'][50]['stockout_pct']
                print(f"    {d}:  {v:.1f}%")
            elif '門市' in name and '>50%' in name:
                v = r['stores_over_50pct']
                print(f"    {d}:  {v}家 / {r['n_stores']}家")
            elif '門市' in name and '>70%' in name:
                v = r['stores_over_70pct']
                print(f"    {d}:  {v}家 / {r['n_stores']}家")
            elif name.startswith('Top 50 Avg DOH'):
                v = r['tier_data'][50]['avg_doh']
                print(f"    {d}:  {v:.2f} 月")
            elif name.startswith('Top 50 Median WOC'):
                v = r['top50_median_woc_weeks']
                print(f"    {d}:  {v:.1f} 週")
            elif name.startswith('門市平均'):
                v = r['avg_store_top50_shortage']
                print(f"    {d}:  {v:.1f}%")
            elif name.startswith('DOH<1月'):
                v = r['doh_under_1m_pct']
                print(f"    {d}:  {v:.1f}%")
            elif name.startswith('平均可展示'):
                v = r['avg_displayable_brands']
                print(f"    {d}:  {v}")
            elif name.startswith('最差門市'):
                v = r['worst_store_pct']
                print(f"    {d}:  {v:.1f}%")
            elif name.startswith('完全零庫存'):
                v = r['zero_stock_pct']
                print(f"    {d}:  {v:.1f}%")
            elif '平均可展示間數' in name:
                n_tier = 50 if 'Top 50' in name else 100
                v = r['tier_data'][n_tier]['avg_disp_stores']
                print(f"    {d}:  {v:.1f} 間 / {r['n_stores']} 店")
            elif '可展示率' in name:
                n_tier = 50 if 'Top 50' in name else 100
                v = r['tier_data'][n_tier]['avg_disp_ratio']
                print(f"    {d}:  {v:.1f}%")


def generate_excel(all_results, output_path):
    """Generate a comprehensive analysis Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    BLUE = PatternFill('solid', fgColor='4285F4')
    RED_BG = PatternFill('solid', fgColor='EA4335')
    GREEN_BG = PatternFill('solid', fgColor='34A853')
    GRAY_BG = PatternFill('solid', fgColor='F3F3F3')
    WH = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    BB = Font(name='Arial', bold=True, size=10)
    BR = Font(name='Arial', size=10)
    RED_F = Font(name='Arial', bold=True, color='CC0000', size=11)
    CT = Alignment(horizontal='center', vertical='center', wrap_text=True)
    BD = Border(left=Side('thin', 'D9D9D9'), right=Side('thin', 'D9D9D9'),
                top=Side('thin', 'D9D9D9'), bottom=Side('thin', 'D9D9D9'))

    dates = [r['date'] for r in all_results]
    dlabels = [d.strftime('%Y-%m-%d') for d in dates]

    # ── Sheet 1: Dashboard ──
    ws = wb.active
    ws.title = '惡化分析 Dashboard'
    row = 1

    ws.cell(row=row, column=1, value='庫存健康度惡化分析 — Feb → Apr 2026')
    ws.cell(row=row, column=1).font = Font(name='Arial', bold=True, size=14)
    row += 2

    # Key metrics table
    metrics_rows = [
        ('Top 50 庫存不足率', [r['tier_data'][50]['shortage_pct'] for r in all_results], '%'),
        ('Top 50 缺貨率', [r['tier_data'][50]['stockout_pct'] for r in all_results], '%'),
        ('Top 50 緊張率', [r['tier_data'][50]['low_pct'] for r in all_results], '%'),
        ('Top 100 庫存不足率', [r['tier_data'][100]['shortage_pct'] for r in all_results], '%'),
        ('Top 100 缺貨率', [r['tier_data'][100]['stockout_pct'] for r in all_results], '%'),
        ('Top 200 庫存不足率', [r['tier_data'][200]['shortage_pct'] for r in all_results], '%'),
        ('Top 300 庫存不足率', [r['tier_data'][300]['shortage_pct'] for r in all_results], '%'),
        ('', [], ''),  # spacer
        ('Top 50 Avg DOH (月)', [r['tier_data'][50]['avg_doh'] for r in all_results], 'num'),
        ('Top 50 Median DOH (月)', [r['tier_data'][50]['median_doh'] for r in all_results], 'num'),
        ('Top 100 Avg DOH (月)', [r['tier_data'][100]['avg_doh'] for r in all_results], 'num'),
        ('Top 50 Median WOC (週)', [r['top50_median_woc_weeks'] for r in all_results], 'num'),
        ('', [], ''),
        ('門市平均 Top50 不足率', [r['avg_store_top50_shortage'] for r in all_results], '%'),
        ('最差門市 Top50 不足率', [r['worst_store_pct'] for r in all_results], '%'),
        ('最佳門市 Top50 不足率', [r['best_store_pct'] for r in all_results], '%'),
        ('', [], ''),
        ('門市 Top50不足 >50% 家數', [r['stores_over_50pct'] for r in all_results], 'int'),
        ('門市 Top50不足 >60% 家數', [r['stores_over_60pct'] for r in all_results], 'int'),
        ('門市 Top50不足 >70% 家數', [r['stores_over_70pct'] for r in all_results], 'int'),
        ('', [], ''),
        ('Top 50 平均可展示間數', [r['tier_data'][50]['avg_disp_stores'] for r in all_results], 'num'),
        ('Top 50 可展示率(間/總店)', [r['tier_data'][50]['avg_disp_ratio'] for r in all_results], '%'),
        ('Top 100 平均可展示間數', [r['tier_data'][100]['avg_disp_stores'] for r in all_results], 'num'),
        ('Top 100 可展示率(間/總店)', [r['tier_data'][100]['avg_disp_ratio'] for r in all_results], '%'),
        ('', [], ''),
        ('平均可展示品牌數', [r['avg_displayable_brands'] for r in all_results], 'num'),
        ('DOH<1月 SKU占比', [r['doh_under_1m_pct'] for r in all_results], '%'),
        ('完全零庫存 占比', [r['zero_stock_pct'] for r in all_results], '%'),
    ]

    # Headers
    headers = ['指標'] + dlabels + ['Δ (Feb→Latest)']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.fill = BLUE; c.font = WH; c.alignment = CT; c.border = BD
    row += 1

    feb_idx = next((i for i, r in enumerate(all_results) if r['date'].month == 2), 0)
    for name, vals, fmt in metrics_rows:
        if not name:
            row += 1
            continue

        ws.cell(row=row, column=1, value=name).font = BB
        ws.cell(row=row, column=1).border = BD

        for ci, v in enumerate(vals, 2):
            c = ws.cell(row=row, column=ci)
            if fmt == '%':
                c.value = v / 100
                c.number_format = '0.0%'
            elif fmt == 'int':
                c.value = int(v)
            else:
                c.value = round(v, 2)
                c.number_format = '0.00'
            c.font = BR; c.alignment = CT; c.border = BD

        # Delta column
        if len(vals) > 1:
            delta_c = ws.cell(row=row, column=len(vals) + 2)
            delta = vals[-1] - vals[feb_idx]
            if fmt == '%':
                delta_c.value = delta / 100
                delta_c.number_format = '+0.0%;-0.0%'
            elif fmt == 'int':
                delta_c.value = int(delta)
                delta_c.number_format = '+0;-0'
            else:
                delta_c.value = round(delta, 2)
                delta_c.number_format = '+0.00;-0.00'

            # Color: red if worsening
            is_worse = delta > 0 if fmt in ['%', 'int'] else delta < 0
            # For these metrics, LOWER = worse (so flip the color logic)
            lower_is_worse = ['平均可展示品牌數', 'Top 50 Avg DOH (月)', 'Top 50 Median DOH (月)',
                              'Top 100 Avg DOH (月)', 'Top 50 Median WOC (週)', '最佳門市 Top50 不足率',
                              'Top 50 平均可展示間數', 'Top 100 平均可展示間數',
                              'Top 50 可展示率(間/總店)', 'Top 100 可展示率(間/總店)']
            if name in lower_is_worse:
                is_worse = delta < 0
            delta_c.font = Font(name='Arial', bold=True, color='CC0000' if is_worse else '0B8043', size=10)
            delta_c.alignment = CT; delta_c.border = BD

        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 30
    for ci in range(2, len(dlabels) + 3):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    # ── Sheet 2: Cohort tracking (Jan Top50 SKUs) ──
    # This is handled by the trend script; skip here

    # ── Sheet 3: Raw data ──
    ws3 = wb.create_sheet('原始數據')
    raw_headers = ['日期', '門市數', 'SKU數', '總列數',
                   'Overall缺貨%', 'Overall緊張%', 'Overall不足%',
                   'T50缺貨%', 'T50緊張%', 'T50不足%', 'T50 AvgDOH月',
                   'T100缺貨%', 'T100不足%', 'T100 AvgDOH月',
                   'T200不足%', 'T300不足%',
                   '>50%店數', '>60%店數', '>70%店數',
                   '平均品牌數', 'DOH<1月%', '零庫存%',
                   '門市avg不足%', '最差門市%', 'MedianWOC週']
    ws3.append(raw_headers)
    for r in all_results:
        ws3.append([
            r['date'].strftime('%Y-%m-%d'), r['n_stores'], r['unique_skus'], r['total_rows'],
            round(r['overall_stockout_pct'], 2), round(r['overall_low_pct'], 2), round(r['overall_shortage_pct'], 2),
            round(r['tier_data'][50]['stockout_pct'], 2), round(r['tier_data'][50]['low_pct'], 2),
            round(r['tier_data'][50]['shortage_pct'], 2), round(r['tier_data'][50]['avg_doh'], 2),
            round(r['tier_data'][100]['stockout_pct'], 2), round(r['tier_data'][100]['shortage_pct'], 2),
            round(r['tier_data'][100]['avg_doh'], 2),
            round(r['tier_data'][200]['shortage_pct'], 2), round(r['tier_data'][300]['shortage_pct'], 2),
            r['stores_over_50pct'], r['stores_over_60pct'], r['stores_over_70pct'],
            r['avg_displayable_brands'], round(r['doh_under_1m_pct'], 2), round(r['zero_stock_pct'], 2),
            round(r['avg_store_top50_shortage'], 2), round(r['worst_store_pct'], 2),
            round(r['top50_median_woc_weeks'], 2),
        ])
    for ci in range(1, len(raw_headers) + 1):
        c = ws3.cell(row=1, column=ci)
        c.fill = BLUE; c.font = WH; c.alignment = CT

    wb.save(output_path)
    print(f"\n  [OUTPUT] {output_path}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--skip', nargs='*', default=[], help='Dates to skip (YYYY-MM-DD)')
    args = parser.parse_args()

    skip = set(args.skip)
    dates_to_use = [d for d in TARGET_DATES if str(d) not in skip]

    print("=" * 70)
    print("  INVENTORY HEALTH DECLINE ANALYSIS")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  Dates: {', '.join(str(d) for d in dates_to_use)}")
    print("=" * 70)

    # Find files
    file_map = {}
    for d in dates_to_use:
        fname = f'StoreHealth_frames_{SALES_DAYS}d_eday{d}.xlsx'
        fpath = os.path.join(BASE_DIR, fname)
        if os.path.exists(fpath):
            file_map[d] = fpath
            print(f"  [OK] {fname}")
        else:
            print(f"  [MISSING] {fname} — skipping")

    if len(file_map) < 2:
        print("\n  ERROR: Need at least 2 dates to compare!")
        sys.exit(1)

    # Analyze each
    all_results = []
    for d in sorted(file_map.keys()):
        result, _ = load_and_analyze(d, file_map[d])
        all_results.append(result)

    # Summary
    print_summary(all_results)

    # Excel
    out = os.path.join(BASE_DIR, f'Decline_Analysis_{all_results[0]["date"]}~{all_results[-1]["date"]}.xlsx')
    generate_excel(all_results, out)

    print(f"\n{'='*70}")
    print(f"  Done! Output: {os.path.basename(out)}")
    print(f"{'='*70}")


if __name__ == '__main__':
    main()
