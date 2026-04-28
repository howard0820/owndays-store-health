"""
store_health_core.py ─ OWNDAYS Taiwan Store Inventory Health Analysis
Core analysis engine: parse 在庫表, compute DOH, identify stockout/dead stock,
suggest replenishment with cross-store transfer recommendations.

Shared by both interactive and automated scripts.
"""

import pandas as pd
import numpy as np
import os, re, warnings, io
from datetime import datetime
from urllib import request as urllib_request
from urllib.error import URLError
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
warnings.filterwarnings('ignore')

# ═══════════════════════════════════════════════════════════════════════════════
# CONFIG
# ═══════════════════════════════════════════════════════════════════════════════
OOS_THRESHOLD = 1         # qty <= 1 = stockout
LOW_STOCK_DOH = 6.0       # DOH < 6.0 = low stock (in weeks, relative to sales period)
TARGET_DOH = 3.0          # target DOH for replenishment (weeks)
MIN_REPLENISH_QTY = 2     # minimum replenishment qty
DONOR_KEEP_MIN = 2        # donor must keep >= 2 after transfer
DEAD_STOCK_QTY_MULT = 5   # qty > mean * 5 = dead stock
DEAD_STOCK_DOH_MULT = 2   # DOH > store_avg * 2 = dead stock
BRAND_DISPLAYABLE_MIN = 2 # qty >= 2 = displayable SKU

# Category classification
# Note: CL is under 雑貨 in the system, identified by 品番 containing ODCLEARVISION
CATEGORY_MAP = {
    'フレーム':       '光學眼鏡',
    'サングラス':     '太陽眼鏡',
    'コンタクトレンズ': '隱形眼鏡',
}
CL_HINBAN_PATTERN = re.compile(r'ODCLEARVISION', re.IGNORECASE)
CL_TRIAL_PATTERN = re.compile(r'FOR TRIAL', re.IGNORECASE)

# CL product line codes (after ODCLEARVISION prefix)
CL_LINES = {
    'MG': 'MG',
    'NH': 'NH',
    'SB': 'SB',
    'GR': 'GR',
    'BR': 'BR',
}

EXCLUDE_STORE_PATTERNS = [
    '本部', '催事', '補聴器', '在庫確保', 'メーカー返品', '修理品',
    '予備在庫', '海外ロス', '加工在庫', '工場', '未検収', 'コンタクトレンズ',
    '店舗回収', '外部倉庫', 'OpenSesame',
    '事務所', '待報廢', 'E-Commerce', 'MOMO', 'Shopee', 'PChome',
    '商品センター', '芝浦倉庫',
]
WAREHOUSE_PATTERNS = ['台湾倉庫', '台灣倉庫', '上海', 'Shanghai']

# Styling
HEADER_FILL = PatternFill('solid', fgColor='4285F4')
DARK_FILL   = PatternFill('solid', fgColor='434343')
RED_FILL    = PatternFill('solid', fgColor='F4CCCC')
YELLOW_FILL = PatternFill('solid', fgColor='FFF2CC')
GREEN_FILL  = PatternFill('solid', fgColor='D9EAD3')
ORANGE_FILL = PatternFill('solid', fgColor='FCE5CD')
ALT_GRAY    = PatternFill('solid', fgColor='F3F3F3')
TOTAL_BLUE  = PatternFill('solid', fgColor='CFE2F3')
WHITE_BOLD  = Font(name='Arial', bold=True, color='FFFFFF', size=10)
BLACK_BOLD  = Font(name='Arial', bold=True, color='000000', size=10)
BLACK_REG   = Font(name='Arial', bold=False, color='000000', size=10)
RED_BOLD    = Font(name='Arial', bold=True, color='CC0000', size=10)
CENTER      = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT        = Alignment(horizontal='left', vertical='center', wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'))


# ═══════════════════════════════════════════════════════════════════════════════
# UTILITIES
# ═══════════════════════════════════════════════════════════════════════════════

def _repair_xlsx(filepath):
    with open(filepath, 'rb') as f:
        data = f.read()
    eocd = data.rfind(b'PK\x05\x06')
    if eocd < 0:
        return
    comment_len = int.from_bytes(data[eocd+20:eocd+22], 'little')
    valid_end = eocd + 22 + comment_len
    if valid_end < len(data):
        with open(filepath, 'wb') as f:
            f.write(data[:valid_end])


def parse_cl_degree(hinban: str) -> str:
    """Extract contact lens degree from 品番.
    Product lines: MG, NH, SB, GR, BR + bare ODCLEARVISION + degree
    e.g. 'TW ODCLEARVISIONMG2.75' -> '-2.75'
         'TW ODCLEARVISIONGR0.25' -> '-0.25'
         'TW ODCLEARVISION8.00'   -> '-8.00'
    """
    s = hinban.strip()
    # Match: product_line_code + degree at end
    # Product lines: MG, NH, SB, GR, BR (all = minus diopter)
    m = re.search(r'(?:MG|NH|SB|GR|BR)(\d+\.\d+)$', s)
    if m:
        return f"-{m.group(1)}"
    # Bare ODCLEARVISION + degree (no product line prefix)
    m2 = re.search(r'ODCLEARVISION(\d+\.\d+)$', s)
    if m2:
        return f"-{m2.group(1)}"
    # Integer format (rare)
    m3 = re.search(r'(?:MG|NH|SB|GR|BR)(\d{3,4})$', s)
    if m3:
        return f"-{int(m3.group(1))/100:.2f}"
    # 0.00 = PLANO
    if s.endswith('0.00'):
        return "PLANO"
    return "PLANO"


def parse_cl_line(hinban: str) -> str:
    """Extract CL product line code from 品番. e.g. MG, NH, SB, GR, BR or 'STD'."""
    s = hinban.strip()
    m = re.search(r'ODCLEARVISION(MG|NH|SB|GR|BR)', s)
    return m.group(1) if m else 'STD'


def is_retail_store(name) -> bool:
    name = str(name) if name else ''
    if not name or name == 'nan':
        return False
    if any(p in name for p in EXCLUDE_STORE_PATTERNS):
        return False
    if any(p in name for p in WAREHOUSE_PATTERNS):
        return False
    return True


def is_warehouse(name) -> bool:
    name = str(name) if name else ''
    return any(p in name for p in WAREHOUSE_PATTERNS)


# ═══════════════════════════════════════════════════════════════════════════════
# LIVE STORE LIST (Google Sheets)
# ═══════════════════════════════════════════════════════════════════════════════

STORE_LIST_URL = (
    "https://docs.google.com/spreadsheets/d/"
    "1A6Oe3Ie8-HqJ_1-vy8ePAeFFbcNEJ9zPQMCx6BK_o5w"
    "/export?format=csv&gid=173795654"
)

# ── Hardcoded alias: 在庫表 store name → (Google Sheet name, store number) ──
# This is the DEFINITIVE mapping for all known mismatches between the 在庫表
# names and Google Sheet names. Checked first before any fuzzy matching.
# Last updated: 2026-04-29
STORE_ALIAS = {
    # --- Spacing / casing differences ---
    'JC PARK新莊':               ('JC PARK 新莊',         '995'),
    'LaLaport南港':              ('Lalaport南港',          '1051'),
    'LaLaport台中':              ('Lalaport台中',          '1029'),
    '微風 台北車站':               ('微風台北車站',           '917'),
    # --- Substring / abbreviation differences ---
    '京站小碧潭':                 ('京站',                  '923'),
    '台北永康':                   ('永康',                  '1036'),
    '新光影城桃園高鐵':            ('桃園高鐵',              '1002'),
    '新光三越信義A11':             ('新光三越A11',            '964'),
    # --- Name order / format differences ---
    '新光三越台南小西門':           ('台南三越小西門',          '939'),
    '秀泰嘉義':                   ('嘉義秀泰',              '988'),
    '新光三越嘉義':               ('嘉義新光',              '1060'),
    # --- MITSUI / 三井 branding differences ---
    'MITSUI OUTLET PARK 台南':   ('三井OUTLET台南',         '1015'),
    'Mitsui outlet park 林口':   ('三井林口 OUTLET',         '935'),
    # --- 大潤發 → 大全聯 rebrand (same store, new tenant name) ---
    '內湖大全聯':                 ('內湖大潤發',             '1022'),
    '員林大全聯':                 ('員林大潤發',             '1008'),
    '頭份大全聯':                 ('頭份大潤發',             '1040'),
    # --- Micro-style differences ---
    '台北微風廣場':               ('微風台北車站',            '917'),
    # --- 淡水大都會廣場 (#959) vs 淡水 (#1059) — two different stores ---
    '淡水大都會廣場':              ('淡水大都會',             '959'),
    '淡水大都會':                 ('淡水大都會',             '959'),
}
# Note: '大巨蛋SOGO' is a new store not yet in the Google Sheet.
# Howard should add it manually; until then it will be kept-by-sales if active.

# Minimum total sales (frame units, summed over the analysis period) for the
# "kept-by-sales" fallback. Active stores typically do hundreds to thousands
# per 30 days; closed stores leak in via stragglers (returns, transfer-in
# corrections) usually well under 30. Override via env var if needed.
MIN_SALES_FOR_FALLBACK = int(os.environ.get('MIN_SALES_FOR_FALLBACK', '30'))

def fetch_live_store_list(timeout=15):
    """Fetch active store list from Google Sheets.
    Returns (live_dict, not_live_set):
      - live_dict: store_name -> store_number for Live stores
      - not_live_set: set of store names explicitly marked Not Live
    Returns (None, set()) if fetch fails.
    Sheet columns: A=店鋪No., B=店名, C=分區, D=Status
    """
    try:
        req = urllib_request.Request(STORE_LIST_URL, headers={'User-Agent': 'Mozilla/5.0'})
        resp = urllib_request.urlopen(req, timeout=timeout)
        csv_data = resp.read().decode('utf-8')
        df = pd.read_csv(io.StringIO(csv_data), header=1)  # header at row 2 (0-indexed row 1)

        # Find columns by name
        num_col = None
        name_col = None
        status_col = None
        for c in df.columns:
            col_str = str(c).strip()
            if '店鋪No' in col_str or '店鋪号' in col_str or '店號' in col_str:
                num_col = c
            if '店名' in col_str:
                name_col = c
            if col_str.lower() in ('status', 'live', '狀態'):
                status_col = c

        if name_col is None:
            name_col = df.columns[1] if len(df.columns) > 1 else df.columns[0]
        if num_col is None:
            num_col = df.columns[0]
        # Column D fallback (0-indexed column 3)
        if status_col is None and len(df.columns) > 3:
            status_col = df.columns[3]

        stores = {}
        not_live = set()
        skipped_not_live = 0
        for idx, row in df.iterrows():
            name_val = row[name_col] if name_col in row.index else None
            num_val = row[num_col] if num_col in row.index else None

            if pd.isna(name_val):
                continue
            s = str(name_val).strip()
            if not s or s == 'nan':
                continue

            # Check status column — only include "Live" stores
            if status_col is not None and status_col in row.index:
                status_val = str(row[status_col]).strip().lower()
                if status_val not in ('live', ''):
                    skipped_not_live += 1
                    not_live.add(s)
                    continue

            store_num = ''
            if num_val and not pd.isna(num_val):
                store_num = str(num_val).strip()
            stores[s] = store_num

        print(f"    -> Live store list: {len(stores)} active stores (skipped {skipped_not_live} non-Live: {not_live})")
        return stores, not_live
    except (URLError, Exception) as e:
        print(f"    -> WARN: Could not fetch live store list ({e})")
        print(f"    -> Continuing with all stores from data")
        return None, set()


def filter_by_live_list(retail_stores, live_stores_dict, sales_data=None, not_live_names=None):
    """Filter retail stores to only those in the live list (dict: store_name -> store_number).
    Uses fuzzy matching. Stores explicitly marked Not Live are always excluded.
    Returns (matched_stores_list, store_number_map).
    """
    if not live_stores_dict:
        return retail_stores, {}

    if not_live_names is None:
        not_live_names = set()

    matched = []
    removed = []
    store_number_map = {}

    alias_hits = 0
    for store in retail_stores:
        # 0) Hardcoded alias lookup (highest priority, guaranteed correct)
        if store in STORE_ALIAS:
            _gs_name, _gs_num = STORE_ALIAS[store]
            matched.append(store)
            store_number_map[store] = _gs_num
            alias_hits += 1
            continue

        # 1) Exact match against Google Sheet names
        if store in live_stores_dict:
            matched.append(store)
            store_number_map[store] = live_stores_dict[store]
            continue

        # 2) Fuzzy matching: substring with length ratio >= 45%
        found = False
        for ls, store_num in live_stores_dict.items():
            if ls in store or store in ls:
                shorter = min(len(ls), len(store))
                longer = max(len(ls), len(store))
                if shorter >= longer * 0.45:
                    matched.append(store)
                    store_number_map[store] = store_num
                    found = True
                    break

        if not found:
            # Second-pass: more aggressive matching — substring match without length ratio,
            # or match after stripping common prefixes (OWNDAYS, OWNDAYS )
            for ls, store_num in live_stores_dict.items():
                # Try matching core name (strip OWNDAYS prefix from both sides)
                store_clean = re.sub(r'^OWNDAYS\s*', '', store)
                ls_clean = re.sub(r'^OWNDAYS\s*', '', ls)
                if store_clean and ls_clean and (ls_clean in store_clean or store_clean in ls_clean):
                    matched.append(store)
                    store_number_map[store] = store_num
                    found = True
                    print(f"    -> MATCHED (aggressive): {store} ↔ {ls} → {store_num}")
                    break

        if not found:
            # Check if explicitly Not Live — NEVER keep these regardless of sales
            is_not_live = store in not_live_names
            if not is_not_live:
                # Also check fuzzy match against not_live_names
                for nl in not_live_names:
                    if nl in store or store in nl:
                        is_not_live = True
                        break
            if is_not_live:
                store_sales = sales_data.get(store, 0) if sales_data is not None else 0
                removed.append((store, store_sales))
                print(f"    -> EXCLUDED (Not Live): {store}")
                continue

            # Fallback: keep only if sales >= threshold (filters out stragglers
            # from closed stores while preserving active stores whose names
            # don't quite match the Live list).
            store_sales = sales_data.get(store, 0) if sales_data is not None else 0
            if store_sales >= MIN_SALES_FOR_FALLBACK:
                matched.append(store)
                store_number_map[store] = ''
                print(f"    -> KEPT (sales={store_sales:.0f} >= {MIN_SALES_FOR_FALLBACK}, but not in live list): {store}")
            else:
                removed.append((store, store_sales))

    if removed:
        print(f"    -> Excluded {len(removed)} closed/unlisted stores (sales < {MIN_SALES_FOR_FALLBACK}):")
        for s, sales in sorted(removed, key=lambda x: -x[1]):
            print(f"       - {s} (sales={sales:.0f})")

    kept_by_sales = len([s for s in matched if store_number_map.get(s)==''])
    matched_with_num = len(matched) - kept_by_sales
    print(f"    -> Active stores: {len(matched)} (alias={alias_hits}, matched={matched_with_num}, kept-by-sales={kept_by_sales})")
    return matched, store_number_map


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 1: PARSE 在庫表
# ═══════════════════════════════════════════════════════════════════════════════

def parse_inventory_file(filepath):
    """Parse 在庫表 Excel into a clean DataFrame with inventory + sales per store × SKU."""
    print(f"  Parsing: {os.path.basename(filepath)}")
    _repair_xlsx(filepath)

    df_raw = pd.read_excel(filepath, header=None, dtype=str, nrows=10)
    header_row = None
    for i in range(min(10, len(df_raw))):
        row0 = str(df_raw.iloc[i, 0]).strip() if pd.notna(df_raw.iloc[i, 0]) else ''
        if row0 == '店舗名' or (pd.notna(df_raw.iloc[i, 1]) and '商品区分CD' in str(df_raw.iloc[i, 1])):
            header_row = i
            break
        if '商品区分CD' in row0:
            header_row = i
            break
    if header_row is None:
        header_row = 5

    df = pd.read_excel(filepath, header=header_row, dtype=str)

    col_map = {}
    for c in df.columns:
        cl = str(c).strip()
        if cl == '店舗名':       col_map[c] = 'store_name'
        elif cl == '商品区分':    col_map[c] = '商品区分'
        elif cl == '中分類' and 'CD' not in cl: col_map[c] = '中分類'
        elif cl == '類別':       col_map[c] = '類別'
        elif cl == 'ブランド':    col_map[c] = 'ブランド'
        elif cl == '品番':       col_map[c] = '品番'
        elif cl in ('カラー2', 'カラー'): col_map[c] = 'カラー'
        elif cl == 'PLU':        col_map[c] = 'PLU'
        elif cl == '商品ステータス': col_map[c] = 'ステータス'
        elif cl == '区分':       col_map[c] = 'kubun'
        elif cl == '数量':       col_map[c] = 'qty'
    df.rename(columns=col_map, inplace=True)

    for req in ['store_name', '品番', 'PLU', 'kubun', 'qty']:
        if req not in df.columns:
            raise ValueError(f"Missing column '{req}'. Found: {list(df.columns)}")

    df['qty'] = pd.to_numeric(df['qty'], errors='coerce').fillna(0).astype(int)
    df['PLU'] = df['PLU'].astype(str).str.strip()
    df['品番'] = df['品番'].astype(str).str.strip()
    if 'カラー' in df.columns:
        df['カラー'] = df['カラー'].astype(str).str.strip()
    else:
        df['カラー'] = ''
    if 'ブランド' not in df.columns:
        df['ブランド'] = ''
    else:
        df['ブランド'] = df['ブランド'].astype(str).str.strip()
    if '商品区分' not in df.columns:
        df['商品区分'] = ''
    else:
        df['商品区分'] = df['商品区分'].astype(str).str.strip()

    # Fill NaN in key columns BEFORE groupby (CL rows often have NaN in 類別/ステータス/store_name)
    for col in ['store_name', '商品区分', '中分類', '類別', 'ブランド', '品番', 'カラー', 'PLU', 'ステータス']:
        if col in df.columns:
            df[col] = df[col].fillna('')

    # Drop rows with empty store_name (orphan rows)
    df = df[df['store_name'].str.strip() != ''].copy()

    inv_df = df[df['kubun'].str.contains('在庫', na=False)].copy()
    sales_df = df[df['kubun'].str.contains('売上', na=False)].copy()

    product_cols = [c for c in ['商品区分', '中分類', '類別', 'ブランド', '品番', 'カラー', 'PLU', 'ステータス']
                    if c in inv_df.columns]
    group_cols = ['store_name'] + product_cols

    inv_agg = inv_df.groupby(group_cols, as_index=False)['qty'].sum()
    inv_agg.rename(columns={'qty': 'inventory'}, inplace=True)

    sales_agg = sales_df.groupby(['store_name', 'PLU'], as_index=False)['qty'].sum()
    sales_agg.rename(columns={'qty': 'sales'}, inplace=True)

    merged = inv_agg.merge(sales_agg, on=['store_name', 'PLU'], how='left')
    merged['sales'] = merged['sales'].fillna(0).astype(int)

    print(f"    -> {len(merged):,} rows, {merged['PLU'].nunique()} SKUs, "
          f"{merged['store_name'].nunique()} stores")
    return merged


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 2: CLASSIFY & COMPUTE
# ═══════════════════════════════════════════════════════════════════════════════

def classify_category(df):
    """Add 'category' column: 光學眼鏡 / 太陽眼鏡 / 隱形眼鏡 / 其他
    CL is under 雑貨 in the system, identified by ODCLEARVISION in 品番.
    TRIAL lenses are excluded.
    """
    df['category'] = df['商品区分'].map(CATEGORY_MAP).fillna('其他')
    # Reclassify ODCLEARVISION items from 雑貨 to 隱形眼鏡
    cl_mask = df['品番'].str.contains(CL_HINBAN_PATTERN, na=False)
    trial_mask = df['品番'].str.contains(CL_TRIAL_PATTERN, na=False)
    df.loc[cl_mask & ~trial_mask, 'category'] = '隱形眼鏡'
    return df


def add_cl_degree(df):
    """For contact lens rows, parse degree and product line from 品番."""
    cl_mask = df['category'] == '隱形眼鏡'
    df.loc[cl_mask, 'degree'] = df.loc[cl_mask, '品番'].apply(parse_cl_degree)
    df.loc[cl_mask, 'cl_line'] = df.loc[cl_mask, '品番'].apply(parse_cl_line)
    df['degree'] = df['degree'].fillna('')
    df['cl_line'] = df['cl_line'].fillna('')
    return df


def compute_metrics(df, sales_days, cl_sales_days=60):
    """Compute DOH (days of hand), daily velocity, and flags per row."""
    # Daily velocity for frames/sunglasses
    frame_mask = df['category'].isin(['光學眼鏡', '太陽眼鏡'])
    cl_mask = df['category'] == '隱形眼鏡'

    df['daily_velocity'] = 0.0
    if sales_days > 0:
        df.loc[frame_mask, 'daily_velocity'] = df.loc[frame_mask, 'sales'] / sales_days
    if cl_sales_days > 0:
        df.loc[cl_mask, 'daily_velocity'] = df.loc[cl_mask, 'sales'] / cl_sales_days

    # DOH in days (relative to the sales period used)
    df['doh_days'] = np.where(
        df['daily_velocity'] > 0,
        df['inventory'] / df['daily_velocity'],
        np.where(df['inventory'] > 0, 9999, 0)
    )
    # DOH in weeks for easier interpretation
    df['doh_weeks'] = df['doh_days'] / 7
    df['doh_months'] = df['doh_days'] / 30

    # Flags
    df['is_stockout'] = df['inventory'] <= OOS_THRESHOLD
    df['is_low_stock'] = (~df['is_stockout']) & (df['doh_weeks'] < LOW_STOCK_DOH)
    df['needs_replenish'] = df['is_stockout'] | df['is_low_stock']

    return df


def compute_rankings(df, sales_days, category_groups=None):
    """Compute national + per-store SKU rankings by sales.
    Rankings are computed WITHIN each category group so that
    frames and CL don't compete for the same rank numbers.

    category_groups: dict mapping group_label -> list of category names.
        If None, ranks across all categories together (legacy behavior).
    """
    if category_groups is None:
        # Legacy: rank everything together
        national = df.groupby('PLU', as_index=False).agg(
            national_sales=('sales', 'sum'),
            national_inv=('inventory', 'sum'),
        )
        national['national_rank'] = national['national_sales'].rank(ascending=False, method='min').astype(int)
        df = df.merge(national[['PLU', 'national_sales', 'national_rank']], on='PLU', how='left')
        df['store_rank'] = df.groupby('store_name')['sales'].rank(ascending=False, method='min').astype(int)
        return df

    # Rank within each category group independently
    df['national_sales'] = np.nan
    df['national_rank'] = np.nan
    df['store_rank'] = np.nan

    for group_label, cats in category_groups.items():
        mask = df['category'].isin(cats)
        sub = df.loc[mask].copy()
        if sub.empty:
            continue

        # National ranking within this group
        national = sub.groupby('PLU', as_index=False).agg(
            national_sales=('sales', 'sum'),
            national_inv=('inventory', 'sum'),
        )
        national['national_rank'] = national['national_sales'].rank(ascending=False, method='min').astype(int)
        sub = sub.merge(national[['PLU', 'national_sales', 'national_rank']], on='PLU', how='left',
                        suffixes=('_old', ''))
        # Drop old columns if they exist from merge
        for col in ['national_sales_old', 'national_rank_old']:
            if col in sub.columns:
                sub.drop(columns=[col], inplace=True)

        # Per-store ranking within this group
        sub['store_rank'] = sub.groupby('store_name')['sales'].rank(ascending=False, method='min').astype(int)

        # Write back
        df.loc[mask, 'national_sales'] = sub['national_sales'].values
        df.loc[mask, 'national_rank'] = sub['national_rank'].values
        df.loc[mask, 'store_rank'] = sub['store_rank'].values

    df['national_rank'] = df['national_rank'].fillna(0).astype(int)
    df['store_rank'] = df['store_rank'].fillna(0).astype(int)

    return df


def compute_replenishment(df, sales_days):
    """Compute suggested replenishment qty = max(target_DOH_gap, MIN_REPLENISH_QTY) - current_inv."""
    target_days = TARGET_DOH * 7  # convert weeks to days
    df['target_inv'] = np.ceil(df['daily_velocity'] * target_days).astype(int)
    df['target_inv'] = df[['target_inv']].apply(lambda x: max(x.iloc[0], MIN_REPLENISH_QTY), axis=1)
    df['replenish_qty'] = (df['target_inv'] - df['inventory']).clip(lower=0).astype(int)
    # Only replenish items that need it
    df.loc[~df['needs_replenish'], 'replenish_qty'] = 0
    return df


def find_transfer_donors(df, store_name, plu, needed_qty):
    """Find candidate donor stores for a specific SKU transfer.
    Donor criteria:
    - Different retail store
    - High DOH for this SKU
    - After transfer, still has >= DONOR_KEEP_MIN qty
    Returns list of dicts sorted by DOH desc.
    """
    candidates = df[
        (df['PLU'] == plu) &
        (df['store_name'] != store_name) &
        (df['store_name'].apply(is_retail_store)) &
        (df['inventory'] > DONOR_KEEP_MIN) &
        (df['doh_weeks'] > TARGET_DOH)
    ].copy()

    if candidates.empty:
        return []

    candidates['available'] = candidates['inventory'] - DONOR_KEEP_MIN
    candidates = candidates[candidates['available'] > 0]
    candidates = candidates.sort_values('doh_weeks', ascending=False)

    results = []
    remaining = needed_qty
    for _, row in candidates.iterrows():
        if remaining <= 0:
            break
        give = min(int(row['available']), remaining)
        results.append({
            'donor_store': row['store_name'],
            'give_qty': give,
            'donor_inv': int(row['inventory']),
            'donor_doh': round(row['doh_weeks'], 1),
            'donor_sales': int(row['sales']),
        })
        remaining -= give

    return results


def batch_transfer_suggestions(df_all, needs_df, max_donors=3):
    """Pre-compute transfer donors for all needs at once (much faster than per-row calls).
    Returns dict: (store_name, PLU) -> donor_string
    """
    if needs_df.empty:
        return {}

    # Build donor pool once: all retail stores with excess DOH
    retail_mask = df_all['store_name'].apply(is_retail_store)
    donor_pool = df_all[
        retail_mask &
        (df_all['inventory'] > DONOR_KEEP_MIN) &
        (df_all['doh_weeks'] > TARGET_DOH)
    ][['store_name', 'PLU', 'inventory', 'doh_weeks', 'sales']].copy()
    donor_pool['available'] = donor_pool['inventory'] - DONOR_KEEP_MIN

    # Group by PLU for fast lookup
    donor_by_plu = {}
    for plu, grp in donor_pool.groupby('PLU'):
        donor_by_plu[plu] = grp.sort_values('doh_weeks', ascending=False)

    result = {}
    for _, row in needs_df.iterrows():
        store = row['store_name']
        plu = row['PLU']
        needed = int(row.get('replenish_qty', 2))

        donors = donor_by_plu.get(plu, pd.DataFrame())
        if donors.empty:
            result[(store, plu)] = ''
            continue

        # Exclude self
        donors = donors[donors['store_name'] != store]
        parts = []
        remaining = needed
        for _, d in donors.head(max_donors * 2).iterrows():
            if remaining <= 0:
                break
            give = min(int(d['available']), remaining)
            if give > 0:
                parts.append(f"{d['store_name']}(給{give}個,DOH={d['doh_weeks']/4.333:.1f}m)")
                remaining -= give
            if len(parts) >= max_donors:
                break

        result[(store, plu)] = '; '.join(parts)

    return result


def identify_dead_stock(df):
    """Identify dead stock: qty abnormally high or DOH abnormally high."""
    frame_mask = df['category'].isin(['光學眼鏡', '太陽眼鏡'])
    retail_mask = df['store_name'].apply(is_retail_store)
    active = df[frame_mask & retail_mask].copy()

    if active.empty:
        df['is_dead_stock'] = False
        df['dead_reason'] = ''
        return df

    # Per-store mean qty and mean DOH
    store_stats = active.groupby('store_name').agg(
        mean_qty=('inventory', 'mean'),
        mean_doh=('doh_weeks', lambda x: x[x < 9999].mean() if (x < 9999).any() else 0),
    )
    active = active.merge(store_stats, on='store_name', how='left')

    # Flag dead stock
    active['is_dead_qty'] = active['inventory'] > (active['mean_qty'] * DEAD_STOCK_QTY_MULT)
    active['is_dead_doh'] = (active['doh_weeks'] < 9999) & (active['doh_weeks'] > active['mean_doh'] * DEAD_STOCK_DOH_MULT)
    active['is_dead_stock'] = active['is_dead_qty'] | active['is_dead_doh']

    reasons = []
    for _, row in active.iterrows():
        r = []
        if row['is_dead_qty']:
            r.append(f"qty={int(row['inventory'])} >> avg={row['mean_qty']:.1f}")
        if row['is_dead_doh']:
            r.append(f"DOH={row['doh_weeks']/4.333:.1f}m >> avg={row['mean_doh']/4.333:.1f}m")
        reasons.append('; '.join(r))
    active['dead_reason'] = reasons

    # Merge back
    df = df.merge(
        active[['store_name', 'PLU', 'is_dead_stock', 'dead_reason']],
        on=['store_name', 'PLU'], how='left', suffixes=('', '_dead')
    )
    df['is_dead_stock'] = df['is_dead_stock'].fillna(False)
    df['dead_reason'] = df['dead_reason'].fillna('')
    return df


def brand_display_analysis(df, sales_days):
    """Per store × brand analysis for display status."""
    retail_mask = df['store_name'].apply(is_retail_store)
    frame_mask = df['category'].isin(['光學眼鏡', '太陽眼鏡'])
    active = df[retail_mask & frame_mask].copy()

    if active.empty:
        return pd.DataFrame()

    active['is_displayable'] = active['inventory'] >= BRAND_DISPLAYABLE_MIN

    brand_store = active.groupby(['store_name', 'ブランド']).agg(
        total_sales=('sales', 'sum'),
        total_inv=('inventory', 'sum'),
        sku_count=('PLU', 'nunique'),
        displayable_sku_count=('is_displayable', 'sum'),
        displayable_inv=('inventory', lambda x: x[active.loc[x.index, 'is_displayable']].sum()
                         if active.loc[x.index, 'is_displayable'].any() else 0),
    ).reset_index()

    # Sales % per store
    store_total = brand_store.groupby('store_name')['total_sales'].transform('sum')
    brand_store['sales_pct'] = np.where(store_total > 0, brand_store['total_sales'] / store_total, 0)

    # DOH for displayable inventory
    brand_store['daily_vel'] = np.where(
        sales_days > 0, brand_store['total_sales'] / sales_days, 0)
    brand_store['doh_weeks'] = np.where(
        brand_store['daily_vel'] > 0,
        brand_store['total_inv'] / brand_store['daily_vel'] / 7,
        np.where(brand_store['total_inv'] > 0, 9999, 0)
    )
    brand_store['doh_months'] = brand_store['doh_weeks'] / 4.333

    brand_store = brand_store.sort_values(['store_name', 'total_sales'], ascending=[True, False])
    return brand_store


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN ANALYSIS PIPELINE
# ═══════════════════════════════════════════════════════════════════════════════

def run_analysis(frame_file, cl_file, sales_days, cl_sales_days=60,
                 store_filter=None, top_n_stores=None, base_date=None):
    """
    Run full store health analysis.

    Args:
        frame_file: path to 在庫表 for frames/sunglasses (with sales_days period)
        cl_file: path to 在庫表 for contact lenses (with cl_sales_days period), or None
        sales_days: sales period in days for frames
        cl_sales_days: sales period for CL (default 60)
        store_filter: list of store names/numbers to include, or None for all
        top_n_stores: if set, only include top N stores by total sales

    Returns:
        dict with analysis results
    """
    print("\n" + "=" * 60)
    print("  Store Health Analysis")
    print("=" * 60)

    # Parse frame/sunglasses file
    print(f"\n[1] Parsing frame/sunglasses file (sales period: {sales_days}d)...")
    df_frame = parse_inventory_file(frame_file)
    df_frame = classify_category(df_frame)

    # Keep only frames and sunglasses from main file
    df_main = df_frame[df_frame['category'].isin(['光學眼鏡', '太陽眼鏡'])].copy()

    # Parse CL file if provided
    df_cl = pd.DataFrame()
    if cl_file and os.path.exists(cl_file):
        print(f"\n[2] Parsing contact lens file (sales period: {cl_sales_days}d)...")
        df_cl_raw = parse_inventory_file(cl_file)
        df_cl_raw = classify_category(df_cl_raw)
        df_cl = df_cl_raw[df_cl_raw['category'] == '隱形眼鏡'].copy()
        if not df_cl.empty:
            df_cl = add_cl_degree(df_cl)
            print(f"    -> CL rows: {len(df_cl):,}, degrees: {df_cl['degree'].nunique()}")
    elif cl_file is None:
        # Try to get CL from the same file
        df_cl = df_frame[df_frame['category'] == '隱形眼鏡'].copy()
        if not df_cl.empty:
            df_cl = add_cl_degree(df_cl)
            print(f"    -> CL from same file: {len(df_cl):,} rows")

    # Combine
    df_all = pd.concat([df_main, df_cl], ignore_index=True)
    if df_all.empty:
        print("ERROR: No data found!")
        return None

    # Filter stores
    retail_stores = [s for s in df_all['store_name'].unique() if is_retail_store(s)]
    wh_stores = [s for s in df_all['store_name'].unique() if is_warehouse(s)]
    print(f"\n[3] Stores: {len(retail_stores)} retail, {len(wh_stores)} warehouse")

    # Cross-check with live store list from Google Sheets
    print(f"\n    Fetching live store list from Google Sheets...")
    live_stores_dict, not_live_names = fetch_live_store_list()
    store_number_map = {}
    # Build per-store sales for fallback logic
    store_sales_totals = df_all.groupby('store_name')['sales'].sum().to_dict()
    if live_stores_dict:
        retail_stores, store_number_map = filter_by_live_list(
            retail_stores, live_stores_dict, sales_data=store_sales_totals,
            not_live_names=not_live_names)

    if store_filter:
        retail_stores = [s for s in retail_stores
                         if any(f in s for f in store_filter)]
        print(f"    -> Filtered to {len(retail_stores)} stores")

    if top_n_stores:
        store_sales = df_all[df_all['store_name'].isin(retail_stores)].groupby('store_name')['sales'].sum()
        top_stores = store_sales.nlargest(top_n_stores).index.tolist()
        retail_stores = top_stores
        print(f"    -> Top {top_n_stores} stores selected")

    # Filter to retail + warehouse
    df_retail = df_all[df_all['store_name'].isin(retail_stores + wh_stores)].copy()

    # Compute metrics
    print(f"\n[4] Computing DOH and flags...")
    df_retail = compute_metrics(df_retail, sales_days, cl_sales_days)
    df_retail = compute_rankings(df_retail, sales_days, category_groups={
        'frames': ['光學眼鏡', '太陽眼鏡'],
        'cl': ['隱形眼鏡'],
    })
    df_retail = compute_replenishment(df_retail, sales_days)
    df_retail = identify_dead_stock(df_retail)

    # Brand analysis
    print(f"\n[5] Brand display analysis...")
    brand_df = brand_display_analysis(df_retail, sales_days)

    # Separate frames and CL
    df_frames = df_retail[
        df_retail['category'].isin(['光學眼鏡', '太陽眼鏡']) &
        df_retail['store_name'].isin(retail_stores)
    ].copy()

    df_contact = df_retail[
        (df_retail['category'] == '隱形眼鏡') &
        df_retail['store_name'].isin(retail_stores)
    ].copy()

    # Summary stats
    print(f"\n[6] Building summary...")
    summary = {
        'total_retail_stores': len(retail_stores),
        'total_frame_skus': df_frames['PLU'].nunique(),
        'total_cl_skus': df_contact['PLU'].nunique() if not df_contact.empty else 0,
        'frame_stockout_pct': (df_frames['is_stockout'].sum() / len(df_frames) * 100) if len(df_frames) > 0 else 0,
        'frame_low_stock_pct': (df_frames['is_low_stock'].sum() / len(df_frames) * 100) if len(df_frames) > 0 else 0,
        'cl_stockout_pct': (df_contact['is_stockout'].sum() / len(df_contact) * 100) if len(df_contact) > 0 else 0,
        'dead_stock_count': df_frames['is_dead_stock'].sum() if 'is_dead_stock' in df_frames.columns else 0,
        'sales_days': sales_days,
        'cl_sales_days': cl_sales_days,
        'base_date': str(base_date) if base_date else '',
    }

    print(f"\n  Summary:")
    print(f"    Retail stores: {summary['total_retail_stores']}")
    print(f"    Frame SKUs: {summary['total_frame_skus']}")
    print(f"    CL SKUs: {summary['total_cl_skus']}")
    print(f"    Frame stockout: {summary['frame_stockout_pct']:.1f}%")
    print(f"    Frame low stock: {summary['frame_low_stock_pct']:.1f}%")
    print(f"    Dead stock items: {summary['dead_stock_count']}")

    return {
        'df_all': df_retail,
        'df_frames': df_frames,
        'df_contact': df_contact,
        'brand_df': brand_df,
        'retail_stores': retail_stores,
        'store_number_map': store_number_map,
        'summary': summary,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 3: GENERATE EXCEL REPORT
# ═══════════════════════════════════════════════════════════════════════════════

def _style_header(ws, row, cols, fill=None, font=None):
    if fill is None: fill = HEADER_FILL
    if font is None: font = WHITE_BOLD
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _write_row(ws, row, data, font=None, fill=None, num_fmt=None):
    if font is None: font = BLACK_REG
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = font
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill
        if num_fmt and isinstance(val, (int, float)):
            cell.number_format = num_fmt


def _auto_width(ws, min_width=8, max_width=30):
    for col in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[col_letter].width = max_len


def generate_report(results, output_path):
    """Generate comprehensive Excel report."""
    print(f"\n[7] Generating Excel report...")

    wb = Workbook()
    df_frames = results['df_frames']
    df_contact = results['df_contact']
    brand_df = results['brand_df']
    df_all = results['df_all']

    # Pre-compute all transfer suggestions at once (much faster)
    all_frame_needs = df_frames[df_frames['needs_replenish']].copy()
    print(f"    -> Pre-computing transfer donors for {len(all_frame_needs)} items...")
    donor_map = batch_transfer_suggestions(df_all, all_frame_needs)
    retail_stores = results['retail_stores']
    summary = results['summary']
    sales_days = summary['sales_days']
    cl_sales_days = summary['cl_sales_days']
    base_date_str = summary.get('base_date', '')
    # Compute date labels for headers
    if base_date_str:
        from datetime import timedelta
        bd = datetime.strptime(base_date_str, '%Y-%m-%d').date()
        frame_start = bd - timedelta(days=sales_days - 1)
        cl_start = bd - timedelta(days=cl_sales_days - 1)
        sales_label = f'全台銷量\n({frame_start}~{bd})'
        inv_label = f'全台庫存\n({bd})'
        cl_sales_label = f'全台銷量\n({cl_start}~{bd})'
        cl_inv_label = f'全台庫存\n({bd})'
    else:
        sales_label = f'全台銷量({sales_days}d)'
        inv_label = '全台庫存'
        cl_sales_label = f'全台銷量({cl_sales_days}d)'
        cl_inv_label = '全台庫存'

    # ── Sheet 1: 總覽 Summary ─────────────────────────────────────────────
    ws = wb.active
    ws.title = '總覽'
    ws.sheet_properties.tabColor = '4285F4'

    headers = ['指標', '數值']
    ws.append(headers)
    _style_header(ws, 1, 2)

    metrics = [
        ('分析門市數', summary['total_retail_stores']),
        ('鏡框/太陽眼鏡 SKU數', summary['total_frame_skus']),
        ('隱形眼鏡 SKU數', summary['total_cl_skus']),
        (f'鏡框銷售期間', f"{summary['sales_days']}天"),
        (f'隱形眼鏡銷售期間', f"{summary['cl_sales_days']}天"),
        ('鏡框缺貨率 (qty≤1)', f"{summary['frame_stockout_pct']:.1f}%"),
        ('鏡框庫存緊張率 (DOH<1.5月)', f"{summary['frame_low_stock_pct']:.1f}%"),
        ('隱形眼鏡缺貨率', f"{summary['cl_stockout_pct']:.1f}%"),
        ('無用庫存筆數', int(summary['dead_stock_count'])),
    ]
    for r, (k, v) in enumerate(metrics, 2):
        ws.cell(row=r, column=1, value=k).font = BLACK_BOLD
        ws.cell(row=r, column=1).alignment = LEFT
        ws.cell(row=r, column=2, value=v).font = BLACK_REG
        ws.cell(row=r, column=2).alignment = CENTER

    # ── Top 10 Dead Stock Items (national quick reference) ────────
    start_row = len(metrics) + 3
    if 'is_dead_stock' in df_frames.columns:
        dead_frames = df_frames[df_frames['is_dead_stock'] == True].copy()
    else:
        dead_frames = pd.DataFrame()

    if not dead_frames.empty:
        dead_national = dead_frames.groupby('PLU', as_index=False).agg(
            品番=('品番', 'first'),
            ブランド=('ブランド', 'first') if 'ブランド' in dead_frames.columns else ('品番', 'first'),
            total_inv=('inventory', 'sum'),
            store_count=('store_name', 'nunique'),
        ).nlargest(10, 'total_inv')

        if not dead_national.empty:
            ws.cell(row=start_row, column=1, value='⚫ 全國無用庫存 Top 10').font = Font(name='Arial', bold=True, size=11, color='434343')
            start_row += 1

            dead_headers = ['排名', '品番', 'ブランド', '總庫存', '涉及門市數']
            for ci, h in enumerate(dead_headers, 1):
                ws.cell(row=start_row, column=ci, value=h)
            _style_header(ws, start_row, len(dead_headers), fill=DARK_FILL)
            start_row += 1

            for idx, (_, row) in enumerate(dead_national.iterrows(), 1):
                data = [
                    idx, row['品番'], row.get('ブランド', ''),
                    int(row['total_inv']), int(row['store_count']),
                ]
                _write_row(ws, start_row, data, fill=ALT_GRAY)
                start_row += 1
            start_row += 1

    # ── Per-store Top N 庫存不足% summary ────────────────────────
    start_row += 1
    ws.cell(row=start_row, column=1, value='📊 各門市庫存不足% (缺貨 + 庫存緊張)').font = Font(name='Arial', bold=True, size=12, color='1A73E8')
    start_row += 1

    store_number_map = results.get('store_number_map', {})
    store_summary_headers = ['No.', '店號', '門市', 'Top 50', 'Top 100', 'Top 200', 'Top 300']
    for ci, h in enumerate(store_summary_headers, 1):
        ws.cell(row=start_row, column=ci, value=h)
    _style_header(ws, start_row, len(store_summary_headers))
    start_row += 1

    # Helper: compute top-N shortage % for a given frame subset
    # per_store=False (national): filter by national_rank <= N (each SKU has rows across all stores)
    # per_store=True: use .head(N) since each store has one row per SKU
    def _top_n_shortage(frames_subset, per_store=False):
        row_data = {}
        sorted_frames = frames_subset.sort_values('national_rank', ascending=True)
        for top_n in [50, 100, 200, 300]:
            if per_store:
                tier = sorted_frames.head(top_n)
            else:
                tier = sorted_frames[sorted_frames['national_rank'] <= top_n]
            n_actual = len(tier)
            if n_actual == 0:
                row_data[top_n] = 'N/A'
            else:
                n_short = int(tier['is_stockout'].sum()) + int(tier['is_low_stock'].sum())
                pct = n_short / n_actual * 100
                row_data[top_n] = f'{pct:.1f}%'
        return row_data

    # National row first
    national_pcts = _top_n_shortage(df_frames)
    _write_row(ws, start_row, ['—', '—', '全國',
               national_pcts[50], national_pcts[100], national_pcts[200], national_pcts[300]],
               fill=PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid'))
    ws.cell(row=start_row, column=3).font = BLACK_BOLD
    start_row += 1

    # Per-store rows sorted by Top 50 shortage % descending (worst first)
    store_pct_list = []
    for store in sorted(retail_stores):
        sf = df_frames[df_frames['store_name'] == store]
        pcts = _top_n_shortage(sf, per_store=True)
        # Parse top50 % for sorting
        try:
            sort_val = float(pcts[50].replace('%', ''))
        except (ValueError, AttributeError):
            sort_val = -1
        store_pct_list.append((store, pcts, sort_val))

    store_pct_list.sort(key=lambda x: x[2], reverse=True)

    for idx, (store, pcts, sort_val) in enumerate(store_pct_list, 1):
        fill = RED_FILL if sort_val > 70 else (YELLOW_FILL if sort_val > 50 else None)
        store_num = store_number_map.get(store, '')
        _write_row(ws, start_row, [idx, store_num, store, pcts[50], pcts[100], pcts[200], pcts[300]], fill=fill)
        start_row += 1

    # ── Brand Displayable SKU Matrix ────────────────────────────────────
    start_row += 2
    ws.cell(row=start_row, column=1,
            value='🟢 各門市 Brand 可展示SKU數 (在庫≥2)').font = Font(name='Arial', bold=True, size=12, color='0B8043')
    start_row += 1

    if not brand_df.empty:
        # Pivot: store × brand → displayable_sku_count
        brand_pivot = brand_df.pivot_table(
            index='store_name',
            columns='ブランド',
            values='displayable_sku_count',
            aggfunc='sum',
            fill_value=0,
        )
        # Sort brands by total displayable SKUs descending
        brand_order = brand_pivot.sum().sort_values(ascending=False).index.tolist()
        brand_pivot = brand_pivot[brand_order]

        # Sort stores by same order as shortage table (worst first)
        store_order = [s for s, _, _ in store_pct_list]
        brand_pivot = brand_pivot.reindex(store_order).fillna(0).astype(int)

        # Add Total per store
        brand_pivot['Total'] = brand_pivot[brand_order].sum(axis=1)

        # Headers: No., 店號, 門市, Total, Brand1, Brand2, ...
        brand_headers = ['No.', '店號', '門市', 'Total'] + brand_order
        for ci, h in enumerate(brand_headers, 1):
            ws.cell(row=start_row, column=ci, value=h)
        _style_header(ws, start_row, len(brand_headers),
                      fill=PatternFill('solid', fgColor='0B8043'))
        start_row += 1

        # Data rows
        for idx, store in enumerate(store_order, 1):
            if store in brand_pivot.index:
                total_val = int(brand_pivot.loc[store, 'Total'])
                row_vals = brand_pivot.loc[store, brand_order].tolist()
            else:
                total_val = 0
                row_vals = [0] * len(brand_order)
            store_num = store_number_map.get(store, '')
            data = [idx, store_num, store, total_val] + row_vals
            _write_row(ws, start_row, data)
            # Bold the Total column
            ws.cell(row=start_row, column=4).font = BLACK_BOLD
            # Color code brand columns (starting at column 4)
            for bi, val in enumerate(row_vals):
                cell = ws.cell(row=start_row, column=bi + 4)
                if val == 0:
                    cell.font = Font(name='Arial', color='CC0000')
                elif val <= 2:
                    cell.font = Font(name='Arial', color='E67C00')
                else:
                    cell.font = Font(name='Arial', color='0B8043')
            start_row += 1

        # Total row
        grand_total = int(brand_pivot['Total'].sum())
        total_data = ['—', '—', '全國合計', grand_total] + [int(brand_pivot[b].sum()) for b in brand_order]
        _write_row(ws, start_row, total_data,
                   fill=PatternFill('solid', fgColor='E8F0FE'))
        ws.cell(row=start_row, column=3).font = BLACK_BOLD
        ws.cell(row=start_row, column=4).font = BLACK_BOLD
        start_row += 1

    _auto_width(ws)

    # ── Sheet: 全國 (National) ──────────────────────────────────────────
    ws_nat = wb.create_sheet(title='全國')
    ws_nat.sheet_properties.tabColor = '1A73E8'
    current_row = 1

    # Aggregate frames nationally by PLU
    nat_frames = df_frames.groupby('PLU', as_index=False).agg(
        品番=('品番', 'first'),
        カラー=('カラー', 'first') if 'カラー' in df_frames.columns else ('品番', 'first'),
        ブランド=('ブランド', 'first') if 'ブランド' in df_frames.columns else ('品番', 'first'),
        sales=('sales', 'sum'),
        inventory=('inventory', 'sum'),
        national_rank=('national_rank', 'first'),
        is_stockout_stores=('is_stockout', 'sum'),
        is_low_stock_stores=('is_low_stock', 'sum'),
        total_stores=('store_name', 'nunique'),
    )
    nat_frames['daily_velocity'] = nat_frames['sales'] / sales_days if sales_days > 0 else 0
    nat_frames['doh_months'] = np.where(
        nat_frames['daily_velocity'] > 0,
        nat_frames['inventory'] / nat_frames['daily_velocity'] / 30,
        np.where(nat_frames['inventory'] > 0, 9999, 0)
    )
    nat_frames = nat_frames.sort_values('national_rank', ascending=True)

    # National summary header
    ws_nat.cell(row=current_row, column=1, value='📊 全國庫存健康摘要').font = Font(name='Arial', bold=True, size=12, color='1A73E8')
    current_row += 1

    tier_headers = ['排名區間', 'SKU數', '缺貨店數合計', '庫存緊張店數合計', '平均每SKU缺貨店比例']
    for ci, h in enumerate(tier_headers, 1):
        ws_nat.cell(row=current_row, column=ci, value=h)
    _style_header(ws_nat, current_row, len(tier_headers))
    current_row += 1

    n_retail = len(retail_stores)
    for top_n in [50, 100, 200, 300]:
        tier = nat_frames.head(top_n)
        n_actual = len(tier)
        if n_actual == 0:
            continue
        oos_stores = int(tier['is_stockout_stores'].sum())
        low_stores = int(tier['is_low_stock_stores'].sum())
        # Average shortage rate = total shortage instances / (SKUs * stores)
        avg_pct = (oos_stores + low_stores) / (n_actual * n_retail) * 100 if n_retail > 0 else 0

        label = f'Top {top_n}' if n_actual >= top_n else f'Top {top_n} (實際{n_actual})'
        fill = RED_FILL if avg_pct > 50 else (YELLOW_FILL if avg_pct > 30 else GREEN_FILL)
        _write_row(ws_nat, current_row,
                   [label, n_actual, oos_stores, low_stores, f'{avg_pct:.1f}%'],
                   fill=fill)
        current_row += 1

    current_row += 1

    # National frame list
    ws_nat.cell(row=current_row, column=1, value='📦 全國鏡框 / 太陽眼鏡 SKU 清單 (按全台銷量排名)').font = Font(name='Arial', bold=True, size=12, color='1A73E8')
    current_row += 1

    nat_frame_headers = ['全台排名', '品番', 'カラー', 'PLU', 'ブランド',
                         sales_label, inv_label, 'DOH(月)',
                         '缺貨店數', '庫存緊張店數', '鋪貨店數']
    for ci, h in enumerate(nat_frame_headers, 1):
        ws_nat.cell(row=current_row, column=ci, value=h)
    _style_header(ws_nat, current_row, len(nat_frame_headers))
    current_row += 1

    for _, row in nat_frames.iterrows():
        doh_display = f"{row['doh_months']:.1f}" if row['doh_months'] < 9999 else '∞'
        oos_cnt = int(row['is_stockout_stores'])
        low_cnt = int(row['is_low_stock_stores'])

        if oos_cnt > n_retail * 0.5:
            fill = RED_FILL
        elif (oos_cnt + low_cnt) > n_retail * 0.3:
            fill = YELLOW_FILL
        else:
            fill = None

        data = [
            int(row['national_rank']), row['品番'], row.get('カラー', ''), row['PLU'],
            row.get('ブランド', ''),
            int(row['sales']), int(row['inventory']), doh_display,
            oos_cnt, low_cnt, int(row['total_stores']),
        ]
        _write_row(ws_nat, current_row, data, fill=fill)
        current_row += 1

    current_row += 2

    # National CL list
    if not df_contact.empty:
        nat_cl = df_contact.groupby('PLU', as_index=False).agg(
            品番=('品番', 'first'),
            degree=('degree', 'first') if 'degree' in df_contact.columns else ('品番', 'first'),
            sales=('sales', 'sum'),
            inventory=('inventory', 'sum'),
            national_rank=('national_rank', 'first'),
            is_stockout_stores=('is_stockout', 'sum'),
            is_low_stock_stores=('is_low_stock', 'sum'),
            total_stores=('store_name', 'nunique'),
        )
        nat_cl['daily_velocity'] = nat_cl['sales'] / cl_sales_days if cl_sales_days > 0 else 0
        nat_cl['doh_months'] = np.where(
            nat_cl['daily_velocity'] > 0,
            nat_cl['inventory'] / nat_cl['daily_velocity'] / 30,
            np.where(nat_cl['inventory'] > 0, 9999, 0)
        )
        nat_cl = nat_cl.sort_values('sales', ascending=False)

        ws_nat.cell(row=current_row, column=1, value='🟡 全國隱形眼鏡 SKU 清單').font = Font(name='Arial', bold=True, size=12, color='E67C00')
        current_row += 1

        nat_cl_headers = ['排名', '品番', '度數', 'PLU', cl_sales_label, cl_inv_label,
                          'DOH(月)', '缺貨店數', '庫存緊張店數', '鋪貨店數']
        for ci, h in enumerate(nat_cl_headers, 1):
            ws_nat.cell(row=current_row, column=ci, value=h)
        _style_header(ws_nat, current_row, len(nat_cl_headers))
        current_row += 1

        for rank_idx, (_, row) in enumerate(nat_cl.iterrows(), 1):
            doh_display = f"{row['doh_months']:.1f}" if row['doh_months'] < 9999 else '∞'
            oos_cnt = int(row['is_stockout_stores'])

            fill = RED_FILL if oos_cnt > n_retail * 0.5 else None

            data = [
                rank_idx, row['品番'], row.get('degree', ''), row['PLU'],
                int(row['sales']), int(row['inventory']), doh_display,
                oos_cnt, int(row['is_low_stock_stores']), int(row['total_stores']),
            ]
            _write_row(ws_nat, current_row, data, fill=fill)
            current_row += 1

    _auto_width(ws_nat, min_width=10, max_width=40)

    # ── Per-store sheets ──────────────────────────────────────────────────
    for store in sorted(retail_stores):
        store_frames = df_frames[df_frames['store_name'] == store].copy()
        store_cl = df_contact[df_contact['store_name'] == store].copy() if not df_contact.empty else pd.DataFrame()
        store_brand = brand_df[brand_df['store_name'] == store].copy() if not brand_df.empty else pd.DataFrame()

        # Truncate store name for sheet title (max 31 chars)
        sheet_name = store[:28].replace('/', '-').replace('\\', '-').replace('*', '').replace('?', '').replace('[', '').replace(']', '')
        ws = wb.create_sheet(title=sheet_name)
        ws.sheet_properties.tabColor = 'FF6D01'

        current_row = 1

        # ── Summary: Top N 庫存不足% ─────────────────────────
        # Sort by national_rank so Top N = "全台銷售排名前 N 的 SKU 在這家店的狀況"
        all_frames_by_rank = store_frames.sort_values('national_rank', ascending=True)
        all_frames_sorted = store_frames.sort_values('sales', ascending=False)
        store_num = store_number_map.get(store, '')
        store_header = f'{store} ({store_num})' if store_num else store
        ws.cell(row=current_row, column=1, value=f'📊 {store_header} 庫存健康摘要').font = Font(name='Arial', bold=True, size=12, color='1A73E8')
        current_row += 1

        tier_headers = ['排名區間', 'SKU數', '缺貨數', '庫存緊張數', '庫存不足%']
        for ci, h in enumerate(tier_headers, 1):
            ws.cell(row=current_row, column=ci, value=h)
        _style_header(ws, current_row, len(tier_headers))
        current_row += 1

        for top_n in [50, 100, 200, 300]:
            tier = all_frames_by_rank.head(top_n)
            n_actual = len(tier)
            if n_actual == 0:
                continue
            n_stockout = int(tier['is_stockout'].sum())
            n_low = int(tier['is_low_stock'].sum())
            n_short = n_stockout + n_low
            pct = n_short / n_actual * 100

            label = f'Top {top_n}' if n_actual >= top_n else f'Top {top_n} (實際{n_actual})'
            fill = RED_FILL if pct > 50 else (YELLOW_FILL if pct > 30 else GREEN_FILL)
            _write_row(ws, current_row,
                       [label, n_actual, n_stockout, n_low, f'{pct:.1f}%'],
                       fill=fill)
            current_row += 1

        current_row += 1

        # ── Section A: 鏡框全SKU清單 (sorted by sales desc) ──────
        ws.cell(row=current_row, column=1, value=f'📦 鏡框 / 太陽眼鏡 全SKU清單').font = Font(name='Arial', bold=True, size=12, color='1A73E8')
        current_row += 1

        frame_headers = ['品番', 'カラー', 'PLU', 'ブランド', '銷量', '在庫', 'DOH(月)',
                         '全台排名', '店內排名', '狀態', '建議補貨', '建議調貨門市']
        for ci, h in enumerate(frame_headers, 1):
            ws.cell(row=current_row, column=ci, value=h)
        _style_header(ws, current_row, len(frame_headers))
        current_row += 1

        for _, row in all_frames_sorted.iterrows():
            if row['is_stockout']:
                status = '缺貨'
                fill = RED_FILL
            elif row['is_low_stock']:
                status = '庫存緊張'
                fill = YELLOW_FILL
            else:
                status = '貨量OK'
                fill = None

            doh_display = f"{row['doh_months']:.1f}" if row['doh_months'] < 9999 else '∞'
            donor_str = donor_map.get((store, row['PLU']), '') if row['needs_replenish'] else ''
            replenish = int(row['replenish_qty']) if row['needs_replenish'] else ''

            data = [
                row['品番'], row.get('カラー', ''), row['PLU'], row.get('ブランド', ''),
                int(row['sales']), int(row['inventory']), doh_display,
                int(row['national_rank']), int(row['store_rank']),
                status, replenish, donor_str,
            ]
            _write_row(ws, current_row, data, fill=fill)
            current_row += 1

        if all_frames_sorted.empty:
            ws.cell(row=current_row, column=1, value='（無鏡框資料）').font = Font(name='Arial', italic=True, color='888888')
            current_row += 1

        current_row += 2

        # ── Section B: 隱形眼鏡全SKU清單 ──────────────────
        ws.cell(row=current_row, column=1, value=f'🟡 隱形眼鏡全SKU清單').font = Font(name='Arial', bold=True, size=12, color='E67C00')
        current_row += 1

        cl_headers = ['品番', '度數', 'PLU', '銷量(60d)', '在庫', 'DOH(月)', '狀態', '建議補貨']
        for ci, h in enumerate(cl_headers, 1):
            ws.cell(row=current_row, column=ci, value=h)
        _style_header(ws, current_row, len(cl_headers))
        current_row += 1

        if not store_cl.empty:
            all_cl_sorted = store_cl.sort_values('sales', ascending=False)
            for _, row in all_cl_sorted.iterrows():
                if row['is_stockout']:
                    status = '缺貨'
                    fill = RED_FILL
                elif row['is_low_stock']:
                    status = '庫存緊張'
                    fill = YELLOW_FILL
                else:
                    status = '貨量OK'
                    fill = None

                doh_display = f"{row['doh_months']:.1f}" if row['doh_months'] < 9999 else '∞'
                replenish = int(row['replenish_qty']) if row['needs_replenish'] else ''

                data = [
                    row['品番'], row.get('degree', ''), row['PLU'],
                    int(row['sales']), int(row['inventory']), doh_display,
                    status, replenish,
                ]
                _write_row(ws, current_row, data, fill=fill)
                current_row += 1
        else:
            ws.cell(row=current_row, column=1, value='（無隱形眼鏡資料）').font = Font(name='Arial', italic=True, color='888888')
            current_row += 1

        current_row += 2

        # ── Section C: Brand可陳列分析 ────────────────────────
        ws.cell(row=current_row, column=1, value=f'🟢 Brand 可陳列分析').font = Font(name='Arial', bold=True, size=12, color='0B8043')
        current_row += 1

        brand_headers = ['品牌', '佔銷量%', 'SKU數', '可陳列SKU(≥2)', '總庫存', '可銷庫存', 'DOH(月)']
        for ci, h in enumerate(brand_headers, 1):
            ws.cell(row=current_row, column=ci, value=h)
        _style_header(ws, current_row, len(brand_headers), fill=PatternFill('solid', fgColor='0B8043'))
        current_row += 1

        if not store_brand.empty:
            for _, row in store_brand.iterrows():
                doh_display = f"{row['doh_months']:.1f}" if row['doh_months'] < 9999 else '∞'
                data = [
                    row['ブランド'],
                    f"{row['sales_pct']:.1%}",
                    int(row['sku_count']),
                    int(row['displayable_sku_count']),
                    int(row['total_inv']),
                    int(row['displayable_inv']),
                    doh_display,
                ]
                _write_row(ws, current_row, data)
                current_row += 1
        else:
            ws.cell(row=current_row, column=1, value='（無品牌資料）').font = Font(name='Arial', italic=True, color='888888')
            current_row += 1

        current_row += 2

        # ── Section D: 無用庫存 ───────────────────────────────
        ws.cell(row=current_row, column=1, value=f'⚫ 無用庫存 (Dead Stock)').font = Font(name='Arial', bold=True, size=12, color='434343')
        current_row += 1

        dead_headers = ['品番', 'カラー', 'PLU', 'ブランド', '銷量', '在庫', 'DOH(月)', '原因']
        for ci, h in enumerate(dead_headers, 1):
            ws.cell(row=current_row, column=ci, value=h)
        _style_header(ws, current_row, len(dead_headers), fill=DARK_FILL)
        current_row += 1

        if 'is_dead_stock' in store_frames.columns:
            dead = store_frames[store_frames['is_dead_stock'] == True].sort_values('inventory', ascending=False)
        else:
            dead = pd.DataFrame()

        for _, row in dead.iterrows():
            doh_display = f"{row['doh_months']:.1f}" if row['doh_months'] < 9999 else '∞'
            data = [
                row['品番'], row.get('カラー', ''), row['PLU'], row.get('ブランド', ''),
                int(row['sales']), int(row['inventory']), doh_display,
                row.get('dead_reason', ''),
            ]
            _write_row(ws, current_row, data, fill=ALT_GRAY)
            current_row += 1

        if dead.empty:
            ws.cell(row=current_row, column=1, value='（無明顯無用庫存）').font = Font(name='Arial', italic=True, color='888888')
            current_row += 1

        _auto_width(ws, min_width=10, max_width=40)

    # ── Sheet: 全台鏡框需求彙總 ──────────────────────────────────────────
    ws_summary = wb.create_sheet(title='全台鏡框補貨需求')
    ws_summary.sheet_properties.tabColor = 'CC0000'

    all_needs = df_frames[df_frames['needs_replenish']].sort_values(
        ['store_name', 'sales'], ascending=[True, False])

    sum_headers = ['店號', '門市', '品番', 'カラー', 'PLU', 'ブランド', '銷量', '在庫',
                   'DOH(月)', '全台排名', '狀態', '建議補貨', '建議調貨門市']
    ws_summary.append(sum_headers)
    _style_header(ws_summary, 1, len(sum_headers))

    row_num = 2
    for _, row in all_needs.iterrows():
        status = '缺貨' if row['is_stockout'] else '庫存緊張'
        doh_display = f"{row['doh_months']:.1f}" if row['doh_months'] < 9999 else '∞'
        donor_str = donor_map.get((row['store_name'], row['PLU']), '')
        store_num = store_number_map.get(row['store_name'], '')

        data = [
            store_num, row['store_name'], row['品番'], row.get('カラー', ''), row['PLU'],
            row.get('ブランド', ''), int(row['sales']), int(row['inventory']),
            doh_display, int(row['national_rank']), status,
            int(row['replenish_qty']), donor_str,
        ]
        fill = RED_FILL if row['is_stockout'] else YELLOW_FILL
        _write_row(ws_summary, row_num, data, fill=fill)
        row_num += 1

    _auto_width(ws_summary, min_width=10, max_width=35)

    # ── Sheet: 全台隱形眼鏡補貨需求 ─────────────────────────────────────────────
    if not df_contact.empty:
        ws_cl = wb.create_sheet(title='全台隱形眼鏡補貨需求')
        ws_cl.sheet_properties.tabColor = 'E67C00'

        cl_needs = df_contact[df_contact['needs_replenish']].sort_values(
            ['store_name', 'sales'], ascending=[True, False])

        cl_sum_headers = ['店號', '門市', '品番', '度數', 'PLU', '銷量(60d)', '在庫',
                          'DOH(月)', '狀態', '建議補貨']
        ws_cl.append(cl_sum_headers)
        _style_header(ws_cl, 1, len(cl_sum_headers))

        row_num = 2
        for _, row in cl_needs.iterrows():
            status = '缺貨' if row['is_stockout'] else '庫存緊張'
            doh_display = f"{row['doh_months']:.1f}" if row['doh_months'] < 9999 else '∞'
            store_num = store_number_map.get(row['store_name'], '')
            data = [
                store_num, row['store_name'], row['品番'], row.get('degree', ''), row['PLU'],
                int(row['sales']), int(row['inventory']), doh_display,
                status, int(row['replenish_qty']),
            ]
            fill = RED_FILL if row['is_stockout'] else YELLOW_FILL
            _write_row(ws_cl, row_num, data, fill=fill)
            row_num += 1
        _auto_width(ws_cl, min_width=10, max_width=35)

    # ── Sheet: 全台Dead Stock ─────────────────────────────────────────────
    ws_dead = wb.create_sheet(title='全台無用庫存')
    ws_dead.sheet_properties.tabColor = '434343'

    if 'is_dead_stock' in df_frames.columns:
        dead_all = df_frames[df_frames['is_dead_stock'] == True].sort_values(
            ['store_name', 'inventory'], ascending=[True, False])
    else:
        dead_all = pd.DataFrame()

    dead_sum_headers = ['店號', '門市', '品番', 'カラー', 'PLU', 'ブランド', '銷量', '在庫', 'DOH(月)', '原因']
    ws_dead.append(dead_sum_headers)
    _style_header(ws_dead, 1, len(dead_sum_headers))

    row_num = 2
    for _, row in dead_all.iterrows():
        doh_display = f"{row['doh_months']:.1f}" if row['doh_months'] < 9999 else '∞'
        store_num = store_number_map.get(row['store_name'], '')
        data = [
            store_num, row['store_name'], row['品番'], row.get('カラー', ''), row['PLU'],
            row.get('ブランド', ''), int(row['sales']), int(row['inventory']),
            doh_display, row.get('dead_reason', ''),
        ]
        _write_row(ws_dead, row_num, data, fill=ALT_GRAY)
        row_num += 1
    _auto_width(ws_dead, min_width=10, max_width=40)

    # ── Sheet: Brand彙總 ──────────────────────────────────────────────────
    if not brand_df.empty:
        ws_brand = wb.create_sheet(title='全台Brand分析')
        ws_brand.sheet_properties.tabColor = '0B8043'

        brand_headers = ['店號', '門市', '品牌', '佔銷量%', 'SKU數', '可陳列SKU(≥2)',
                         '總庫存', '可銷庫存', 'DOH(月)']
        ws_brand.append(brand_headers)
        _style_header(ws_brand, 1, len(brand_headers), fill=PatternFill('solid', fgColor='0B8043'))

        row_num = 2
        for _, row in brand_df.iterrows():
            doh_display = f"{row['doh_months']:.1f}" if row['doh_months'] < 9999 else '∞'
            store_num = store_number_map.get(row['store_name'], '')
            data = [
                store_num, row['store_name'], row['ブランド'],
                f"{row['sales_pct']:.1%}",
                int(row['sku_count']), int(row['displayable_sku_count']),
                int(row['total_inv']), int(row['displayable_inv']),
                doh_display,
            ]
            _write_row(ws_brand, row_num, data)
            row_num += 1
        _auto_width(ws_brand, min_width=10, max_width=30)

    # Save
    wb.save(output_path)
    size = os.path.getsize(output_path)
    print(f"\n  ✓ Report saved: {os.path.basename(output_path)} ({size:,} bytes)")
    print(f"    Sheets: {', '.join(wb.sheetnames)}")
    return output_path
